from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.ingestion.event_resolver import EventMeta
from app.services.event_aggregator import EventReportData, MarketTotals, ParticipantTotals


def _safe_sheet_title(title: str) -> str:
    t = (title or "").strip()
    if not t:
        return "Sheet"
    bad = ['\\', '/', '?', '*', '[', ']', ':']
    for ch in bad:
        t = t.replace(ch, " ")
    return t[:31]


def _set_col_widths(ws, widths: Dict[int, int]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _ts_to_str(ts: int | None) -> str:
    if not ts:
        return ""
    return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def export_event_report_xlsx(
    *,
    event: EventMeta,
    report: EventReportData,
    out_path: str,
) -> str:
    wb = Workbook()
    bold = Font(bold=True)

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"

    ws["A1"], ws["B1"] = "Event slug", report.event_slug
    ws["A2"], ws["B2"] = "Event title", report.event_title
    ws["A3"], ws["B3"] = "Event id", report.event_id
    ws["A4"], ws["B4"] = "As of (UTC)", report.as_of_utc
    ws["A5"], ws["B5"] = "Total trades", report.total_trades
    ws["A6"], ws["B6"] = "Unique traders", report.unique_traders
    ws["A7"], ws["B7"] = "Total turnover (USD)", float(report.total_turnover_usd)

    for r in range(1, 8):
        ws[f"A{r}"].font = bold

    start_row = 9
    headers = [
        "market_id",
        "conditionId",
        "market_slug",
        "question",
        "trades_count",
        "unique_traders",
        "buy_usd",
        "sell_usd",
        "turnover_usd",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = bold

    market_id_by_cid = {m.condition_id: m.market_id for m in event.markets}

    markets_sorted = sorted(report.markets.values(), key=lambda x: x.turnover_usd, reverse=True)
    row = start_row + 1
    for m in markets_sorted:
        ws.cell(row, 1, market_id_by_cid.get(m.condition_id, ""))
        ws.cell(row, 2, m.condition_id)
        ws.cell(row, 3, m.market_slug)
        ws.cell(row, 4, m.question)
        ws.cell(row, 5, m.trades_count)
        ws.cell(row, 6, len(m.unique_traders))
        ws.cell(row, 7, float(m.buy_usd))
        ws.cell(row, 8, float(m.sell_usd))
        ws.cell(row, 9, float(m.turnover_usd))
        row += 1

    ws.freeze_panes = ws["A10"]
    ws.auto_filter.ref = f"A{start_row}:I{row-1}"
    _set_col_widths(ws, {1: 10, 2: 22, 3: 40, 4: 60, 5: 12, 6: 14, 7: 14, 8: 14, 9: 14})

    # --- Per-market sheets ---
    # группируем участников по conditionId
    per_market: Dict[str, list[ParticipantTotals]] = {}
    for (cid, _wallet, _outcome), p in report.participants.items():
        per_market.setdefault(cid, []).append(p)

    for m in markets_sorted:
        title = _safe_sheet_title(market_id_by_cid.get(m.condition_id) or m.market_slug or "market")
        # защита от дублей имён листов
        base = title
        k = 2
        while title in wb.sheetnames:
            title = _safe_sheet_title(f"{base}-{k}")
            k += 1

        wsm = wb.create_sheet(title=title)

        # шапка
        wsm["A1"], wsm["B1"] = "market_slug", m.market_slug
        wsm["A2"], wsm["B2"] = "conditionId", m.condition_id
        wsm["A3"], wsm["B3"] = "question", m.question
        wsm["A4"], wsm["B4"] = "trades_count", m.trades_count
        wsm["A5"], wsm["B5"] = "unique_traders", len(m.unique_traders)
        wsm["A6"], wsm["B6"] = "turnover_usd", float(m.turnover_usd)
        for r in range(1, 7):
            wsm[f"A{r}"].font = bold

        table_row = 8
        p_headers = [
            "trader_name",
            "trader_pseudonym",
            "trader_address",
            "outcome",
            "buy_shares",
            "buy_usd",
            "sell_shares",
            "sell_usd",
            "net_shares",
            "net_spent_usd",
            "avg_buy_price",
            "avg_sell_price",
            "trades_count",
            "first_ts_utc",
            "last_ts_utc",
        ]
        for c, h in enumerate(p_headers, 1):
            cell = wsm.cell(row=table_row, column=c, value=h)
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        plist = per_market.get(m.condition_id, [])
        # сортировка: по net_spent_usd убыв., потом turnover убыв.
        plist.sort(key=lambda p: (p.net_spent_usd, p.buy_usd + p.sell_usd), reverse=True)

        r = table_row + 1
        for p in plist:
            wsm.cell(r, 1, p.trader_name)
            wsm.cell(r, 2, p.trader_pseudonym)
            wsm.cell(r, 3, p.trader_address)
            wsm.cell(r, 4, p.outcome)

            wsm.cell(r, 5, float(p.buy_shares))
            wsm.cell(r, 6, float(p.buy_usd))
            wsm.cell(r, 7, float(p.sell_shares))
            wsm.cell(r, 8, float(p.sell_usd))

            wsm.cell(r, 9, float(p.net_shares))
            wsm.cell(r, 10, float(p.net_spent_usd))

            wsm.cell(r, 11, float(p.avg_buy_price) if p.avg_buy_price is not None else "")
            wsm.cell(r, 12, float(p.avg_sell_price) if p.avg_sell_price is not None else "")

            wsm.cell(r, 13, p.trades_count)
            wsm.cell(r, 14, _ts_to_str(p.first_ts))
            wsm.cell(r, 15, _ts_to_str(p.last_ts))
            r += 1

        wsm.freeze_panes = wsm["A9"]
        wsm.auto_filter.ref = f"A{table_row}:O{max(table_row, r-1)}"
        _set_col_widths(
            wsm,
            {
                1: 18, 2: 18, 3: 44, 4: 8,
                5: 12, 6: 14, 7: 12, 8: 14,
                9: 12, 10: 14, 11: 12, 12: 12,
                13: 12, 14: 20, 15: 20,
            },
        )

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)
